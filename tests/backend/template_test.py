import unittest
from unittest.mock import MagicMock, patch
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from flask import Flask, request, session

import sys
import os

# Get the directory of the current test file
current_dir = os.path.dirname(os.path.abspath(__file__))
# Get the parent directory of 'tests' (which should contain 'app.py')
parent_dir = os.path.dirname(os.path.dirname(current_dir))
# Add the parent directory to sys.path
sys.path.insert(0, parent_dir)

from app import app, normalize_name, REPORT_TEMPLATE_NAME

class TemplateAPITestCase(unittest.TestCase):

    def setUp(self):
        self.app = app.test_client()
        self.app.application.config['TESTING'] = True

        # Mock necessary functions
        self.mock_download = patch('app.download_file_from_file_name', return_value=b'mock_excel_content').start()
        self.mock_upload = patch('app.update_file_from_file_name', return_value=True).start()

        # Create a mock Excel file for testing
        self.mock_wb = Workbook()
        self.mock_ws = self.mock_wb.active
        self.mock_ws['A1'] = 'Template Name'
        self.mock_ws['A2'] = 'test_template'
        self.mock_ws['B2'] = 'Column1'
        self.mock_ws['C2'] = 'Column2'
        self.mock_excel_buffer = BytesIO()
        self.mock_wb.save(self.mock_excel_buffer)
        self.mock_excel_content = self.mock_excel_buffer.getvalue()
        self.mock_download.return_value = self.mock_excel_content

    def tearDown(self):
        self.mock_download.stop()
        self.mock_upload.stop()

    def set_session(self, access_token='mock_token'):
        with self.app.session_transaction() as sess:
            sess['access_token'] = access_token

    def test_update_template_user_made_changes(self):
        self.set_session()
        updated_data = {"name": "test_template", "columns": ["NewColumnA", "NewColumnB", "NewColumnC"]}
        response = self.app.post('/api/update_template', json=updated_data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {"message": "Template updated successfully."})

        # Assert that download and upload were called
        self.mock_download.assert_called_once_with('mock_token', REPORT_TEMPLATE_NAME)
        self.mock_upload.assert_called_once()

        # Verify the content of the uploaded file (basic check)
        args, _ = self.mock_upload.call_args
        uploaded_file_io = args[2]  # Get the BytesIO object
        uploaded_file_io.seek(0)  # Reset the buffer's position
        wb = load_workbook(uploaded_file_io)
        ws = wb.active
        self.assertEqual(normalize_name(ws['A2'].value), 'test_template')
        self.assertEqual(ws['B2'].value, 'NewColumnA')
        self.assertEqual(ws['C2'].value, 'NewColumnB')
        self.assertEqual(ws['D2'].value, 'NewColumnC')
        self.assertIsNone(ws.cell(row=2, column=5).value) # Ensure old columns are cleared

    def test_update_template_user_did_not_make_changes(self):
        self.set_session()
        updated_data = {"name": "test_template", "columns": ["Column1", "Column2"]}
        response = self.app.post('/api/update_template', json=updated_data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {"message": "Template updated successfully."})

        # Assert that download and upload were called
        self.mock_download.assert_called_once_with('mock_token', REPORT_TEMPLATE_NAME)
        self.mock_upload.assert_called_once()

        # Verify the content of the uploaded file (basic check)
        args, _ = self.mock_upload.call_args
        uploaded_file = args[2]  # Get the BytesIO object
        uploaded_file.seek(0)  # Reset the buffer's position
        wb = load_workbook(uploaded_file)
        ws = wb.active
        self.assertEqual(normalize_name(ws['A2'].value), 'test_template')
        self.assertEqual(ws['B2'].value, 'Column1')
        self.assertEqual(ws['C2'].value, 'Column2')
        print(ws.cell(row=2, column=3).value)
        self.assertIsNone(ws.cell(row=2, column=4).value) # Ensure no extra columns

    def test_update_template_upload_fails(self):
        self.set_session()
        self.mock_upload.return_value = False
        updated_data = {"name": "test_template", "columns": ["NewColumnA", "NewColumnB"]}
        response = self.app.post('/api/update_template', json=updated_data)
        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.get_json(), {"error": "Error updating the template. Please try again."})

    def test_delete_template_user_deletes_template(self):
        self.set_session()
        delete_data = {"name": "test_template"}
        response = self.app.post('/api/delete_template', json=delete_data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {"message": "Template deleted successfully."})

        # Assert that download and upload were called
        self.mock_download.assert_called_once_with('mock_token', REPORT_TEMPLATE_NAME)
        self.mock_upload.assert_called_once()

        # Verify the content of the uploaded file (basic check - row should be deleted)
        args, _ = self.mock_upload.call_args
        uploaded_file = args[2]  # Get the BytesIO object
        uploaded_file.seek(0)  # Reset the buffer's position
        wb = load_workbook(uploaded_file)
        ws = wb.active
        self.assertEqual(ws.max_row, 1) # Only the header row should remain

    def test_delete_template_not_found(self):
        self.set_session()
        delete_data = {"name": "non_existent_template"}
        response = self.app.post('/api/delete_template', json=delete_data)
        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.get_json(), {"error": "Template not found in the Excel file."})
        self.mock_upload.assert_not_called()

    def test_delete_template_upload_fails(self):
        self.set_session()
        self.mock_upload.return_value = False
        delete_data = {"name": "test_template"}
        response = self.app.post('/api/delete_template', json=delete_data)
        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.get_json(), {"error": "Error deleting the template. Please try again."})

if __name__ == '__main__':
    unittest.main()