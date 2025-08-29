from flask import Flask, g, session, render_template, jsonify, request, Response, send_file
from dotenv import load_dotenv
import msal
import sqlite3
import webview
import threading
from flask import Flask, render_template, request, session, redirect, url_for, flash
import secrets
import tempfile
from openpyxl import load_workbook, Workbook 
from openpyxl.utils import get_column_letter
from io import BytesIO
import re
from fpdf import FPDF
import os
import json
from datetime import datetime
import time
import pandas as pd
from werkzeug.utils import secure_filename
import requests
import openpyxl
import io
import sys
from werkzeug.exceptions import HTTPException

# Load environment variables
# This ensures it works in a bundled PyInstaller app
load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

# Use correct folder paths for PyInstaller
base_dir = os.path.abspath(os.path.dirname(__file__))

app = Flask(
    __name__,
    template_folder=os.path.join(base_dir, "templates"),
    static_folder=os.path.join(base_dir, "static")
)

# needed for pyinstaller 
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Configuration from environment variables
app.secret_key = os.getenv("SECRET_KEY", secrets.token_hex(16))

STUDENT_DB_URL=os.getenv("STUDENT_DB_URL")
FIELDS_ORDER_URL = os.getenv("FIELDS_ORDER_URL")
STUDENT_FILES_URL = os.getenv("STUDENT_FILES_URL")

REPORT_TEMPLATE_NAME=os.getenv("REPORT_TEMPLATE_NAME")
STUDENT_DB_NAME=os.getenv("STUDENT_DB_NAME")
FIELDS_ORDER_NAME = os.getenv("FIELDS_ORDER_NAME")

STUDENT_DB_LOCAL_PATH="students_local.db"
FIELD_ORDER_LOCAL_PATH="field_order.txt"
global_mode = "view"

# Global paths for templates and field order
TEMPLATES_XLSX_PATH = None
FIELD_ORDER_TXT_PATH = None

# On app startup, check for field_order.txt and templates.xlsx in launch folder
exe_folder = os.path.dirname(os.path.abspath(sys.argv[0]))
if os.path.exists(os.path.join(exe_folder, 'field_order.txt')):
    FIELD_ORDER_TXT_PATH = os.path.join(exe_folder, 'field_order.txt')
if os.path.exists(os.path.join(exe_folder, 'templates.xlsx')):
    TEMPLATES_XLSX_PATH = os.path.join(exe_folder, 'templates.xlsx')

# CATCHES ALL unhandled exceptions and redirects user
@app.errorhandler(Exception)
def handle_exception(e):
    # If it's an HTTP error (like 404), don't flash it
    if isinstance(e, HTTPException):
        return e  # Let Flask handle it normally
    
    flash(f'An error occurred: {str(e)}', 'danger')
    if global_mode:
        return redirect(request.referrer or url_for('database'))
    else:
        # return to index if not logged in
        return redirect(url_for('index')) 
    
#################################################################################
# App specific and routing logic
#################################################################################

@app.route('/exit_app')
def exit_app():
    webview.windows[0].destroy() #closes the window
    return Response(status=204)  # No Content
# logout
@app.route('/logout')
def logout():
    return redirect(url_for('index'))
       
@app.route('/')
def index():
    startup_status = getattr(webview.windows[0], 'startup_status', None) if hasattr(webview, 'windows') and webview.windows else None
    return render_template("login.html", startup_status=startup_status)

@app.route('/database')
def database():
    return render_template("database.html", delete_mode=False)

@app.route('/admin')
def admin():
    return render_template("admin.html")

@app.route('/export_to_excel')
def export_to_excel():
    try:
        
        students = fetch_students() 

        if not students:
            return "No students found"

        #open Save File dialog using PyWebView
        filepath = webview.windows[0].create_file_dialog(webview.SAVE_DIALOG, save_filename='students.xlsx')

        if not filepath:
            return "Export canceled"

        #create Excel file
        wb = Workbook()
        ws = wb.active
        headers = list(students[0].keys())
        ws.append(headers)

        for student in students:
            ws.append(list(student.values()))

        wb.save(filepath)
        return f"File saved to:\n{filepath}"

    except Exception as e:
        return f"Error: {str(e)}"

def fetch_students():
    db = get_db()
    cursor = db.execute("SELECT * FROM students")
    return [dict(row) for row in cursor.fetchall()]


@app.route('/import')
def import_data():
    return render_template('import.html')

def get_color_scheme(id):
    return "default"  # Return the default color scheme

@app.route('/get_color_scheme_session')
def get_color_scheme_session():
    color_scheme = session.get('color_scheme', 'default')
    return jsonify({'color_scheme': color_scheme})

@app.route("/update_color_scheme", methods=["POST"])
def update_color_scheme():
    color_scheme = request.json.get("colorScheme")
    session["color_scheme"] = color_scheme

    return jsonify({"message": "Color scheme updated successfully", "colorScheme": color_scheme})

def get_templates():
    templates_dict = {}
    if TEMPLATES_XLSX_PATH and os.path.exists(TEMPLATES_XLSX_PATH):
        try:
            print(f"Loading templates from {TEMPLATES_XLSX_PATH}")
            wb = load_workbook(TEMPLATES_XLSX_PATH)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):  # Skip the header row
                template_name = row[0].value.upper() if row[0].value else None
                fields = [cell.value.upper() for cell in row[1:] if cell.value]
                if template_name and fields:
                    templates_dict[template_name] = fields
            
            print("Loaded templates:", templates_dict)
            return templates_dict
        except Exception as e:
            print(f"Error reading templates.xlsx: {e}")
            return {}
    else:
        return {}

def get_all_fields(sorted):
    db = get_db()
    cursor = db.execute("PRAGMA table_info(students);")
    fields = [row[1] for row in cursor.fetchall()]
    if "id" in fields:
        fields.remove("id")
    if sorted:
        fields.sort()
    db.close()
    return jsonify(fields)

@app.route('/generate_report', methods=['GET', 'POST'])
def generate_report():

    if request.method == 'GET':
        all_fields = json.loads(get_all_fields(True).data)
        templates_dict = get_templates()

        return render_template('auxiliary/generate_report.html', back_link="/database", templates=templates_dict, all_fields=all_fields)
    if request.method == 'POST':
        student_ids = request.args.getlist('ids[]')

        try:
            # Get JSON data from frontend
            data = request.json
            selected_fields = data.get("fields", [])
            custom_title = data.get("title", "")

            print(student_ids)

            if not selected_fields:
                return jsonify({"success": False, "error": "No fields selected"}), 400
            
            student_data = get_students_by_ids(student_ids, selected_fields)

            # Generate PDF using PyWebView's file dialog
            pdf_path = generate_pdf(student_data, selected_fields, custom_title)

            if not pdf_path:
                return jsonify({"success": False, "error": "User canceled save dialog"}), 400
            try:
                os.startfile(pdf_path)
            except:
                print("can't open it")

            return jsonify({
                "success": True,
                "message": "Report generated successfully",
                "report_path": pdf_path
            })

        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

def generate_pdf(data, fields, custom_title):
    """Generate a PDF report with a cover page and individual student tables on separate pages using FPDF."""
    file_types = ('PDF (*.pdf)', 'All files (*.*)')
    file_path = webview.windows[0].create_file_dialog(
        webview.SAVE_DIALOG,
        file_types=file_types
    )

    if not file_path:
        return None
    if isinstance(file_path, list):
        file_path = file_path[0]

    # Create a PDF object
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_margins(20, 15, 20)

    # ----------------------
    #  Add the cover page
    # ----------------------
    pdf.add_page()
    pdf.set_font("Times", style="B", size=14)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    pdf.cell(0, 10, f"Report generated at: {timestamp}", ln=True, align="C")

    # ----------------------
    #  Add pages per student
    # ----------------------
    for student in data:
        pdf.add_page()

        pdf.set_font("Times", style="B", size=16)
        pdf.cell(0, 10, custom_title, ln=True, align="C")
        pdf.ln(5)

        pdf.set_font("Times", size=12)

        line_height = 10
        col1_width = 60

        for field, value in zip(fields, student):
            # remember where this “row” starts
            x_start = pdf.get_x()
            y_start = pdf.get_y()

            # 1) Field name box (wraps if too long)
            pdf.set_font("Times", style="B", size=12)
            pdf.multi_cell(col1_width, line_height, field, border=1) 
            field_y_end = pdf.get_y()

            # 2) Value box (also wraps)
            pdf.set_xy(x_start + col1_width, y_start)
            pdf.set_font("Times", size=12)
            pdf.multi_cell(0, line_height, str(value), border=1) 
            value_y_end = pdf.get_y()

            # 3) Move down to the greatest of the two box-heights
            y_end = max(field_y_end, value_y_end)
            pdf.set_y(y_end) # Use set_y directly to move to the correct y-coordinate

        pdf.ln(10)

    # Save the PDF
    pdf.output(file_path)
    return file_path

@app.route('/templates')
def templates():
    all_fields = json.loads(get_all_fields(True).data)
    templates_dict = get_templates()
    # Pass the dictionary and error to the template
    return render_template('templates.html', templates_dict=templates_dict, all_fields=all_fields)

@app.route('/new_template', methods=['GET', 'POST'])
def new_template():
    if request.method == 'GET':
        columns = json.loads(get_all_fields(True).data)
        return render_template('auxiliary/new_template.html', back_link="/templates", columns=columns)
    elif request.method == 'POST':
        # Get JSON data from request
        data = request.json
        template_name = data.get("name")
        selected_columns = data.get("columns")

        if not template_name or not selected_columns:
            return {"error": "Template name and columns are required."}, 400

        if not TEMPLATES_XLSX_PATH or not os.path.exists(TEMPLATES_XLSX_PATH):
            return {"error": "templates.xlsx not found."}, 500
        try:
            wb = load_workbook(TEMPLATES_XLSX_PATH)
            ws = wb.active
            next_row = ws.max_row + 1
            ws[f"A{next_row}"] = template_name
            for i, column in enumerate(selected_columns, start=2):
                ws[f"{get_column_letter(i)}{next_row}"] = column
            wb.save(TEMPLATES_XLSX_PATH)
            return {"message": "Template appended successfully."}, 201
        except Exception as e:
            return {"error": f"Error saving template: {str(e)}"}, 500

# Normalize function: convert to lowercase and replace spaces with underscores
def normalize_name(name):
    return re.sub(r"\s+", "_", name.lower()) if name else None

@app.route('/api/update_template', methods=['POST'])
def update_template_api():
    data = request.json
    template_name = data.get("name")  # Template name to update
    updated_columns = data.get("columns")  # New column order

    if not template_name or not updated_columns:
        return {"error": "Template name and updated columns are required."}, 400

    if not TEMPLATES_XLSX_PATH or not os.path.exists(TEMPLATES_XLSX_PATH):
        return {"error": "templates.xlsx not found."}, 500
    try:
        wb = load_workbook(TEMPLATES_XLSX_PATH)
        ws = wb.active

        # Find the row with the given template name
        template_row = None
        normalized_template_name = normalize_name(template_name)

        for row in range(2, ws.max_row + 1):  # Skip header row
            cell_value = normalize_name(ws[f"A{row}"].value)
            if cell_value == normalized_template_name:
                template_row = row
                break

        if not template_row:
            return {"error": "Template not found in the Excel file."}, 404

        # Clear the entire row (except template name in column A)
        for col in range(2, ws.max_column + 1):  # Start from column B
            ws[f"{get_column_letter(col)}{template_row}"] = None

        # Insert new attributes in the cleared row
        for i, column in enumerate(updated_columns, start=2):  # Start at column B
            ws[f"{get_column_letter(i)}{template_row}"] = column

        wb.save(TEMPLATES_XLSX_PATH)
        return {"message": "Template updated successfully."}, 200
    except Exception as e:
        return {"error": f"Error updating template: {str(e)}"}, 500

@app.route('/api/delete_template', methods=['POST'])
def delete_template_api():
    data = request.json
    template_name = data.get("name")  # Template name to delete

    if not template_name:
        return {"error": "Template name is required."}, 400

    if not TEMPLATES_XLSX_PATH or not os.path.exists(TEMPLATES_XLSX_PATH):
        return {"error": "templates.xlsx not found."}, 500
    try:
        wb = load_workbook(TEMPLATES_XLSX_PATH)
        ws = wb.active

        # Find the row with the given template name
        template_row = None
        normalized_template_name = normalize_name(template_name)

        for row in range(2, ws.max_row + 1):  # Skip header row
            cell_value = normalize_name(ws[f"A{row}"].value)
            if cell_value == normalized_template_name:
                template_row = row
                break

        if not template_row:
            return {"error": "Template not found in the Excel file."}, 404

        # Delete the entire row with the template (shift everything up)
        ws.delete_rows(template_row)

        wb.save(TEMPLATES_XLSX_PATH)
        return {"message": "Template deleted successfully."}, 200
    except Exception as e:
        return {"error": f"Error deleting template: {str(e)}"}, 500

@app.route('/layout')
def layout():
    return render_template('auxiliary/layout.html', 
                           heading="", 
                           back_link="/database")

@app.route('/details')
def details():
    return render_template('auxiliary/details.html',
                           heading="",
                           back_link="/database")


# Function to start Flask in a separate thread
def start_flask():
    app.run(port=5000, debug=False)

def get_db():
    if "db" not in g:
        # Check if local copy exists, otherwise create an empty db file
        if not os.path.exists(STUDENT_DB_LOCAL_PATH):
            conn = sqlite3.connect(STUDENT_DB_LOCAL_PATH)
            conn.close()
        g.db = sqlite3.connect(STUDENT_DB_LOCAL_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db

def get_field_order():
    if FIELD_ORDER_TXT_PATH and os.path.exists(FIELD_ORDER_TXT_PATH):
        try:
            with open(FIELD_ORDER_TXT_PATH, "r", encoding="utf-8") as f:
                field_order = f.read().strip().split("\n")
            return [field.strip() for field in field_order if field.strip()]
        except Exception as e:
            print(f"Error reading field order file: {e}")
            return None
    # fallback
    return []

@app.teardown_appcontext
def close_db(exception):
    """Close database connection at the end of request."""
    db = g.pop("db", None)
    if db is not None:
        db.close()

# Route to get all students
@app.route("/students", methods=["GET"])
def get_students():
    db = get_db()
    cursor = db.execute("SELECT * FROM students")
    students = [dict(row) for row in cursor.fetchall()]
    db.close()
    return jsonify(students)

# DATABASE LOGIC
@app.route('/get_fields', methods=['GET'])
def get_fields():
    field_order = get_field_order()
    return jsonify(field_order)

@app.route('/get_data', methods=['POST'])
def get_data():
    data = request.json
    sort = data.get('sort', {})
    filter = data.get('filter', [])
    search = data.get('search', '')
    queried_data = query_db(sort, filter, search)
    return jsonify(queried_data)

def query_db(sort, filter_params, search_term):
    db = get_db()
    field_order = get_field_order()

    # Construct ORDER BY clause
    order_by_clauses = [f'"{field}" {direction}' for field, direction in sort.items()]
    order_by_sql = ", ".join(order_by_clauses) if order_by_clauses else "id ASC"

    # Process filter parameters
    filters = []
    values = []
    for param in filter_params:
        field, value = param.split("-")  # Assuming format is "field-value"
        filters.append(f'"{field}" = ?')
        values.append(value)

    # Add search term filter
    if search_term:
        search_filter = " OR ".join([f'"{field}" LIKE ?' for field in field_order])
        filters.append(f"({search_filter})")
        values.extend([f"%{search_term}%" for _ in field_order])  # Add the search term with wildcards for LIKE clause

    where_clause = " AND ".join(filters) if filters else "1=1"  # Ensure valid WHERE clause

    # Construct the query
    quoted_fields = [f'"{f}"' for f in field_order]
    query = f"SELECT {', '.join(quoted_fields)}, id FROM students WHERE {where_clause} ORDER BY {order_by_sql};"


    students = db.execute(query, values).fetchall()
    result = [dict(row) for row in students]
    db.close()
    return result

@app.route('/get_student_fields', methods=['GET'])
def get_student_fields():
    return get_all_fields(True)

@app.route('/get_student_fields_unsorted', methods=['GET'])
def get_student_fields_unsorted():
    return get_all_fields(False)

@app.route('/get_field_values', methods=['POST'])
def get_field_values():
    data = request.json
    field = data.get('field')
    db = get_db()
    query = f'SELECT DISTINCT "{field}" FROM students;'
    cursor = db.execute(query)
    values = [row[0] for row in cursor.fetchall()]  # Extract values
    db.close()
    return jsonify(values)

@app.route('/get_student', methods=['POST'])
def get_student():
    data = request.get_json() 
    student_id = data.get('id')
    if not student_id:
        return jsonify({"error": "No student ID provided"}), 400
    try:
        student = get_student_by_id(student_id)
        if student:
            return jsonify(student)
        else:
            return jsonify({"error": "Student not found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def get_students_by_ids(ids, selected_fields):
    db = get_db()
    # Fetch student data for selected IDs and fields
    if not selected_fields:  # checks for an empty list
        fields_str = "*"
    else:
        fields_str = ", ".join(f'"{field}"' for field in selected_fields)
    placeholders = ", ".join("?" for _ in ids)
    query = f"SELECT {fields_str} FROM students WHERE id IN ({placeholders})"
    cursor=db.execute(query, ids)
    student_data = cursor.fetchall()
    db.close()
    return student_data

def get_student_by_id(id):
    db = get_db()
    query = "SELECT * FROM students WHERE id = ?"
    cursor = db.execute(query, (id,))
    student = cursor.fetchone()
    db.close()
    if student:
        # Convert row to dictionary using cursor description
        columns = [col[0] for col in cursor.description]
        student_dict = dict(zip(columns, student))
        student_dict.pop('id', None)
        return student_dict
    return None

@app.route('/login', methods=['POST'])
def login():
    if 'excel' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['excel']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'File must be an Excel (.xlsx or .xls)'}), 400
    try:
        wb = load_workbook(file)
        ws = wb.active
        columns = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print(columns)
        db = get_db()
        db.execute('DROP TABLE IF EXISTS students')
        col_defs = ', '.join([f'"{col}" TEXT' for col in columns])
        db.execute(f'CREATE TABLE students ({col_defs})')
        for row in ws.iter_rows(min_row=2, values_only=True):
            placeholders = ','.join(['?'] * len(columns))
            db.execute(f'INSERT INTO students ({','.join(columns)}) VALUES ({placeholders})', row)
        db.commit()
        db.close()
        # If field_order.txt is empty, populate it with columns
        if FIELD_ORDER_TXT_PATH and os.path.exists(FIELD_ORDER_TXT_PATH):
            with open(FIELD_ORDER_TXT_PATH, 'r+', encoding='utf-8') as f:
                content = f.read().strip()
                if not content:
                    f.seek(0)
                    f.write('\n'.join(columns))
                    f.truncate()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': f'Failed to process Excel file: {str(e)}'}), 400

@app.route('/save_fields', methods=['POST'])
def save_fields():
    data = request.json
    fields = data.get('fields')
    if not fields or not isinstance(fields, list):
        return jsonify({'error': 'No fields provided'}), 400
    try:
        with open(FIELD_ORDER_TXT_PATH or 'field_order.txt', 'w', encoding='utf-8') as f:
            f.write('\n'.join(fields))
        return jsonify({'success': True, 'message': 'Field order saved successfully'})
    except Exception as e:
        return jsonify({'error': f'Failed to save fields: {str(e)}'}), 500

if __name__ == '__main__':
    # Scan the folder containing the .exe and show file names
    exe_folder = os.path.dirname(os.path.abspath(sys.argv[0]))
    file_list = os.listdir(exe_folder)
    file_message = "Files in this folder:\n" + "\n".join(file_list)

    startup_status = {
        'templates': 'found' if TEMPLATES_XLSX_PATH else 'created',
        'field_order': 'found' if FIELD_ORDER_TXT_PATH else 'created'
    }

    # Check for templates.xlsx and field_order.txt, create if missing
    if not TEMPLATES_XLSX_PATH:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Template Name", "Field 1", "Field 2", "Field 3"])
        wb.save(os.path.join(exe_folder, 'templates.xlsx'))
        TEMPLATES_XLSX_PATH = os.path.join(exe_folder, 'templates.xlsx')
    if not FIELD_ORDER_TXT_PATH:
        with open(os.path.join(exe_folder, 'field_order.txt'), 'w', encoding='utf-8') as f:
            f.write('')
        FIELD_ORDER_TXT_PATH = os.path.join(exe_folder, 'field_order.txt')

    # Start Flask server in a separate thread
    flask_thread = threading.Thread(target=start_flask)
    flask_thread.daemon = True
    flask_thread.start()

    # Create a PyWebView window to load the Flask app
    window = webview.create_window('SusanDB', 'http://127.0.0.1:5000', frameless=False)
    window.startup_status = startup_status
    webview.start()